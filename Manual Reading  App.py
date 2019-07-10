# -*- coding: utf-8 -*-
"""
Created on Wed Mar 20 11:57:21 2019

@author: devis
"""

__author__ = '{Paolo Veneroso, Devis Meka}'
__version__ = '0.9.0'
__email__ = '{paolo.veneroso@triplenet-energy.com, devis.meka@triplnenet-energy.com}'
__status__ = '{beta}'
##__copyright__ = 'Copyright {year}, {project_name}'
##__credits__ = ['{credit_list}']
##__license__ = '{license}'
##__maintainer__ = '{maintainer}'

from datetime import datetime
from datetime import date
import time
import sys
import requests
from lxml import html
import json
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook
#import last_manual_readings
from colorama import init, Fore
import os
import progressbar
from tkinter import *
from tkinter import filedialog
from tkinter import scrolledtext
import tkinter.ttk as ttk
import numpy
from PIL import Image as it
from PIL import ImageTk as itk
import tkinter as tk
from PIL import Image, ImageTk


def start():
##    image = './image.gif'
##    msg = '''        Welcome to the Manual Readings app!\n
##        Please select an option below.'''
##    title = 'Manual Readings V0.9.0'
    
    global window
    window = tk.Tk()
    window.title("Welcome to Manual Readings app")
    window.geometry('900x500')
    image = Image.open('./newyork.jpg')
    photo = ImageTk.PhotoImage(image)
    ImageLabel = Label(window, image=photo)
    ImageLabel.place(x=0, y=0, relwidth=1, relheight=1)
    

#
   
    global frame1
    frame1 = Frame(window)
    frame1.pack(side=LEFT)
    global frame2
    frame2 = Frame(window)
    frame2.pack(side=LEFT)
    global frame3
    frame3 = Frame(window)
    frame3.pack(side=LEFT)
    global frame4
    frame4 = Frame(window)
    frame4.pack(side=LEFT)
    
#
#    global lb1    
#    lbl = Label(frame1, text='Convert cvs file to Excel')
#    lbl.pack()

    global btn1
    btn1 = Button(frame2, activebackground='grey', width=35, activeforeground='black',text="Convert cvs file to Excel", command=csv_to_excel)
    btn1.pack()
    

#    global lb2
#    lb2 = Label(frame1, text='Send Excel readings to Platform')
#    lb2.pack()

    global btn2
    btn2 = Button(frame2, activebackground='grey', width=35, activeforeground='black',text="Send Excel readings to Platform", command=lambda: building_selection())
    btn2.pack()

#    global lb3
#    lb3 = Label(frame1, text='Exit')
#    lb3.pack()

    global btn3
    btn3 = Button(frame2, activebackground='grey', width=35, activeforeground='black', text="Exit", command=lambda: sys.exit(0))
    btn3.pack()

    global txt
    txt = scrolledtext.ScrolledText(frame3, width=26)
    txt.pack()

    global bar
    bar = ttk.Progressbar(frame3, length=200, orient='horizontal', mode='determinate')
    bar.pack()

    global selection
    selection = Listbox(frame4, width=26, height=21)
  
    window.mainloop()

def csv_to_excel():
    try:
        global btn1
        global bar
        global txt
        
        btn1.config(state=DISABLED)
    
        filepath_in = filedialog.askopenfilename(initialdir = "./",title = "Select the csv file",filetypes = (("csv files","*.csv"),("all files","*.*")))
        if not filepath_in:
            btn1.config(state=NORMAL)
            return None
     
        filepath_out = ".\input.xlsx"
        pd.read_csv(filepath_in, delimiter=";").to_excel(filepath_out, index=False)
    
        workfile = pd.read_excel("input.xlsx")
            
        workfile = workfile.drop(columns=['Breaker','Service Floor','Hub','Panel name','Location','Contract','Max demand','Health','Port','Address','Protocol'],axis=1)
        workfile = workfile.drop(columns=['Timeout','Slave Number','Phase Number','Slot Number','Building Index','Wireless','Lora','CT Size'],axis=1)
        
        txt.insert(INSERT, 'Processing the csv file...')
    
        row_number = len(workfile['TNE number'])
    
        for i in range(0, row_number):
            if str(workfile['Type'][i])!=('Manual') or str(workfile['Status'][i])!=('Online') or str(workfile['Category'][i])!=('Electric'):
                workfile = workfile.drop([i])
    
        row_number = len(workfile['TNE number'])
    
        workfile['estimated']=['false'] * row_number
        workfile['timezone']=["America/New_York"] * row_number
        workfile['kWhTot_multiplier']=[' '] * row_number
        workfile['kW3PhT_multiplier']=[' ']* row_number
        workfile['Last Reading Date']=[' ']* row_number
        workfile['Last Reading']=[' ']* row_number
    
        today = date.today()
        workfile['Reading date']=[str(today.month) + '/' + str(today.day) + '/' + str(today.year)] * row_number
        workfile['kWhTot']=[' '] * row_number
        workfile['kW3PhT']=[' '] * row_number
        workfile['createdBy']=[' '] * row_number
    
        bar["maximum"] = row_number
        current_bar_value = 1
    
        workfile = workfile.reset_index(drop=True)
    
        for i in range(0, row_number):
            word = str(workfile['Meter model'][i])
            result = word.find('mul')
    
            if result == -1:
                workfile['kWhTot_multiplier'][i]='1'
                workfile['kW3PhT_multiplier'][i]='1'
            else:
                result2 = word.find('kW3pht')
                if result2 == -1:
                    workfile['kWhTot_multiplier'][i]=word[result+3:]
                    workfile['kW3PhT_multiplier'][i]=''
                else:
                    workfile['kWhTot_multiplier'][i]=word[result+3:]
                    workfile['kW3PhT_multiplier'][i]=word[result+3:]
    
                last_ts, last_usage = last_manual_readings(workfile['TNE number'][i])
                
                if(last_ts and last_usage):
    ##             print(last_usage)
                    workfile['Last Reading Date'][i]=last_ts
                    workfile['Last Reading'][i]=last_usage
                    
            bar["value"]= current_bar_value
            bar.update()
            time.sleep(0.5)
            current_bar_value = current_bar_value + 1
        
    ##    x1=x1.drop('Unnamed: 0',axis=1)
        workfile.to_excel(filepath_out, sheet_name="Sheet1", index=False)
        
        wb = load_workbook(filepath_out)
        
        wb=load_workbook(filepath_out)
        # get Sheet
        source=wb['Sheet1']
        ws=source
        dv = DataValidation(type="list", formula1='"rmehta,pveneroso2,msuaris"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add('Q2:Q100')
        
        wb.save(filepath_out)
    
        btn1.config(state=NORMAL)
    
        txt.insert(INSERT, '\nDone processing.')
        txt.insert(INSERT, '\nFile input.xlsx created. Please open the file to enter the readings.')
        txt.insert(INSERT, '\n--------------------------\n')
    except IOError :
        txt.insert(INSERT, '\nThe file you trying to convert is open. You must close the file before processing it!')


def last_manual_readings(tne):
    session_requests = requests.session()

    login_url = ''# the real login url was removed
    payload = {'uid':'user',
               'password':'passuser'}

    result = session_requests.post(login_url, data=json.dumps(payload))

    bearer = json.dumps(result.headers["Authorization"])
    bearer = json.loads(bearer)

    headers = {
        'Content-Type': 'application/json',
        'Authorization': bearer
        }

    tne_number = tne

    manual_url = 'http://iuse.triplenet-energy.com/api/meter/' + tne_number + '/lastmanualreadings'

    result_2 = session_requests.get(manual_url, headers=headers)

    meters_json=result_2.content

    meters_json=meters_json.decode('utf-8')
    ##print(meters_json)
    
    if meters_json == '[]':
##        print('TNE Number or manual readings not found!')
        return None, None

    meters_json='{"meters":' + meters_json + '}'
    meters_json=json.loads(meters_json)
    ##print(meters_json)

    try:
        last_ts=''
        last_usage=''
        temp=sorted(meters_json['meters'], key=lambda d: d['ts'], reverse=False)
        temp=temp.pop()
        last_ts=int(temp['ts'])
        last_ts=time.strftime('%m/%d/%Y', time.localtime(last_ts/1000))
##        print('Reading Date: ' + temp_ts + ', kWhTot: '  + str(temp['kWhTot']))
        if(temp['kWhTot']):
            last_usage = temp['kWhTot']
        return last_ts, last_usage
    
    ##    temp=time.strftime('%m/%d/%Y', time.localtime(temp/1000))
    ##    print('Reading Date: ' + temp + ', kWhTot: '  + str(x['kWhTot']))
     
    except (ValueError, KeyError, TypeError):
    ##    print("JSON format error")
##        print('Reading Date: ' + temp_ts)
        return None, None


def building_selection():
##    selection.forget()
####    txt.pack_forget()
####    bar.forget()
##    selection.pack()
    ##    bar.pack()
   # selection = Listbox(frame4, width=26, height=21)
   #selection.pack()

    global btn2
    btn2.config(state=DISABLED)

    global selection
    selection = Listbox(frame4, width=26, height=21)
    selection.insert(1, '1301 Avenue of the Americas')
    selection.insert(2, '1330 Avenue of the Americas')
    selection.insert(3, '1351 Washington Blvd')
    selection.insert(4, '1540 Broadway')
    selection.insert(5, '1633 Broadway')
    selection.insert(6, '31 West 52nd Street')
    selection.insert(7, '340 Madison Avenue')
    selection.insert(8, '50 Beale Street')
    selection.insert(9, '530 Fifth Avenue')
    selection.insert(10, '711 Fifth Avenue')
    selection.insert(11, '745 5th Avenue')
    selection.insert(12, '75 Rockefeller Plaza')
    selection.insert(13, '767 Third Avenue')
    selection.insert(14, '77 Water Street')
    selection.insert(15, '777 Third Avenue')
    selection.insert(16, '809 UN Plaza')
    selection.insert(17, '900 Third Avenue')
    selection.insert(18, 'International Gem Tower')
    selection.insert(19, 'One Front Street')
    selection.insert(20, 'One Market Plaza')
    selection.insert(21, 'The Factory LIC')
    selection.pack()
    
    selection.bind("<Double-1>", excel_to_platform)
    


def excel_to_platform(button):
     
   global selection
   global btn2
   global txt
   global bar

   api = ''
   building = ''
  # try:
   choice = str(selection.curselection()[0] + 1)
   
   if choice == "1":
      build = '1301_6Th_Ave'
      api = '008d9be8-28d7-4fc6-8d8e-2a553352c852'
   elif choice == "2":
      build = '1330_AOFA'
      api = '9473b66c-4a08-11e8-a821-7e90a1d4d9d1'
   elif choice == "3":
      build = '0e4efb56-64ba-46ff-82cb-b4cb59432068'
      api = '90f1e9f5-2932-11e8-9021-7e90a1d4d9d1'
   elif choice == "4":
      build = 'ad1168a4-9bf7-4d11-b5f9-d42046e212ac'
      api = '9d0f79df-69c6-48b4-b7c7-9e8a94785fd3'
   elif choice == "5":
      build = '1633_Broadway'
      api = '517db289-5270-4a66-ae1e-2e9ef5f2e7f8'
   elif choice == "6":
      build = '31_West_52nd_Street'
      api = 'aa5062b4-f496-4c89-87a2-f45fd0cadc93'
   elif choice == "7":
      build = '340_Madison_Avenue'
      api = '27335b6e-0207-4241-9c7c-dd7c2a8c0fae'
   elif choice == "8":
      build = '50_Beale_Street'
      api = '34f634a5-bb60-4fc0-9451-3ef1340633a3'
   elif choice == "9":
      build = '530_Fifth_Avenue'
      api = 'ed409a72-98c9-4a59-ba72-2a5eef1e32c0'
   elif choice == "10":
      build = 'a32fc8b7-21f5-4e8d-be24-aafc420657a3'
      api = 'b129b669-880c-48cb-8582-4c3b9115a9fc'
   elif choice == "11":
      build = '745_5Th_Ave'
      api = 'b5bc36ec-55be-4816-9cb2-84315bfe54ae'
   elif choice == "12":
      build = '75 Rock'
      api = '472d9a9f-237b-11e8-8f65-7e90a1d4d9d1'
   elif choice == "13":
      build = '76821d7d-a469-48a0-a06b-f888ab605af2'
      api = 'a54af631bb0c7946d-1d5a-4c01-865a-38d'
   elif choice == "14":
      build = '77_Water_Street'
      api = '2e4b2a9c-647e-4b48-ae32-4ea38dd49a24'
   elif choice == "15":
      build = 'e36e1381-b087-11e8-919e-02420aff0006'
      api = '1a889c14-bad4-42df-8026-04679f8c975a'
   elif choice == "16":
      build = '809_UN_PLAZA'
      api = 'ddd5259c-2724-49fd-b841-9ef99f93e69c'
   elif choice == "17":
      build = '900_3rd_Ave'
      api = '44342eca-044f-4903-bd60-bafff088f63d'
   elif choice == "18":
      build = 'International_Gem_Tower'
      api = '34a02436-d4c7-47bb-95bf-d4cbf9ad4144'
   elif choice == "19":
      build = '7a34a699-88b2-4f72-a469-a4c791d14f91'
      api = 'cd2d85dc-d199-4540-869e-bfbe616798c7'
   elif choice == "20":
      build = 'One_Market_Plaza'
      api = '06727233-2a0a-4c6c-a8fe-fdd5c99a2e56'
   elif choice == "21":
      build = 'The Factory LIC'
      api = '733c265d-38c0-4e10-a8d2-002248d903e5'

   filepath_in = filedialog.askopenfilename(initialdir = "./",title = "Select the Excel file",filetypes = (("xlsx files","*.xlsx"),("all files","*.*")))

   if not filepath_in:
       btn2.config(state=NORMAL)
       return None


   excel_input = pd.read_excel(filepath_in) # OPEN INPUT FILE
   row_number = (excel_input.shape[0]) # NUMBER OF INPUT ROWS EXCLUDING HEADER

   meters_json = [] # ARRAY OF JSON WITH METERS DATA

   txt.insert(INSERT, 'Processing the xlsx file...')

   bar["maximum"] = row_number
   current_bar_value = 1

   
   # CREATE JSON FILE FOR THE PLATFROM
   for row in range(0,row_number):
      beginning = '{'
      api_key = '"APIKey":"' + api + '", '
      building = '"building":"' + build + '", '

      if str(excel_input.loc[row,'TNE number'])=='nan':
          txt.insert(INSERT, '\nFound error in "TNE number" column, row ' + str(row+2) + '\n')
          return None
      tne_number = '"tne_number":"' + str(excel_input.loc[row,'TNE number']) + '", '

      if str(excel_input.loc[row,'Serial number'])=='nan':
          txt.insert(INSERT, '\nFound error in "Serial number" column, row ' + str(row+2) + '\n')
          return None
      serial_number = '"serial_number":"' + str(excel_input.loc[row,'Serial number']) + '", '

      if str(excel_input.loc[row,'Reading date'])=='nan':
          txt.insert(INSERT, '\nFound error in "Reading date" column, row ' + str(row+2) + '\n')
          return None

##      if type(excel_input.loc[row,'Reading date'])!=pd._libs.tslibs.timestamps.Timestamp and type(excel_input.loc[row,'Reading date'])!=datetime.datetime:
##         print('Found error in "Reading date" column, row ' + str(row+2) + '. Not a date\n')
##         return None

      dt = datetime.strptime(excel_input.loc[row,'Reading date'], '%m/%d/%Y')

      # PARSE THE TIMESTAMP AND CONVERT TO EPOCH
##    dt = excel_input.loc[row,'Reading date']
      ts = str(int(time.mktime(dt.timetuple()))*1000+43200000) # TIME AT 12:00PM
      date = '"ts":' + ts + ', '

      if str(excel_input.loc[row,'timezone'])=='nan':
          txt.insert(INSERT, '\nFound error in "timezone" column, row ' + str(row+2) + '\n')
          return None  
      timezone = '"timezone":"' + str(excel_input.loc[row,'timezone']) + '", '

      if str(excel_input.loc[row,'kWhTot_multiplier'])=='nan':
          txt.insert(INSERT, '\nFound error in "kWhTot_multiplier" column, row ' + str(row+2) + '\n')
          return None

      if type(excel_input.loc[row,'kWhTot_multiplier'])!=numpy.float64 and type(excel_input.loc[row,'kWhTot_multiplier'])!=numpy.int64:
          txt.insert(INSERT, '\nFound error in "kWhTot_multiplier" column, row ' + str(row+2) + '. Not a number\n')
          return None
      kwhtot_mult = '"kWhTot_multiplier":' + str(excel_input.loc[row,'kWhTot_multiplier']) + ', '

      if str(excel_input.loc[row,'kWhTot'])=='nan':
          txt.insert(INSERT, '\nFound error in "kWhTot" column, row ' + str(row+2) + '\n')
          return None
      if type(excel_input.loc[row,'kWhTot'])!=numpy.float64 and type(excel_input.loc[row,'kWhTot'])!=numpy.int64:
          txt.insert(INSERT, '\nFound error in "kWhTot" column, row ' + str(row+2) + '. Not a number\n')
          return None
      kwhtot = '"kWhTot":' + str(excel_input.loc[row,'kWhTot']*excel_input.loc[row,'kWhTot_multiplier']) + ', '

      if str(excel_input.loc[row,'kW3PhT_multiplier'])!='nan': # IF kW MULTIPLIER IS PRESENT
         if type(excel_input.loc[row,'kW3PhT_multiplier'])!=numpy.float64 and type(excel_input.loc[row,'kW3PhT_multiplier'])!=numpy.int64:
            txt.insert(INSERT, '\nFound error in "kW3PhT_multiplier" column, row ' + str(row+2) + '. Not a number\n')
            return None
         kw_mult = '"kW3PhT_multiplier":' + str(excel_input.loc[row,'kW3PhT_multiplier']) + ', '
      else: # SKIPS IF kW MULTIPLIER IS NOT PRESENT
         kw_mult = ''

      if str(excel_input.loc[row,'kW3PhT'])!='nan': # IF kW3PhT IS PRESENT
         if type(excel_input.loc[row,'kW3PhT'])!=numpy.float64 and type(excel_input.loc[row,'kW3PhT'])!=numpy.int64:
            txt.insert(INSERT, '\nFound error in "kW3PhT" column, row ' + str(row+2) + '. Not a number\n')
            return None
         else:
            if kw_mult=='':
               kw = kw = '"kW3PhT":' + str(excel_input.loc[row,'kW3PhT']) + ', '
            else:
               kw = '"kW3PhT":' + str(excel_input.loc[row,'kW3PhT']*excel_input.loc[row,'kW3PhT_multiplier']) + ', '
      else: # SKIP IF kW3PhT IS NOT PRESENT
          kw = ''

      manual = '"manual":true' + ', '

      if excel_input.loc[row,'estimated']==False:
          estimated = '"estimated":false' + ', '
      elif excel_input.loc[row,'estimated']==True:
          estimated = '"estimated":true' + ', '
      else:
          estimated = ''

      if str(excel_input.loc[row,'createdBy'])=='nan':
          txt.insert(INSERT, '\nFound error in "createdBy" column, row ' + str(row+2) + '\n')
          return None
      createdby = '"createdBy":"' + str(excel_input.loc[row,'createdBy']) + '", '

      createdat = '"createdAt":' + str(int(time.time())*1000)
      end = '}'

      output = beginning + api_key + building + tne_number + serial_number + date + timezone + kwhtot + kw + kwhtot_mult + kw_mult + manual + estimated + createdby + createdat + end
      meters_json.append(output)

      bar["value"]= current_bar_value
      bar.update()
      time.sleep(0.5)
      current_bar_value = current_bar_value + 1

    
##   for row in range(0, len(meters_json)):
##      print(meters_json[row])

   url = 'http://iuse.triplenet-energy.com/rest/triplenet/buildingdata'
   headers = {'Content-Type': 'application/json'}

   current_bar_value = 1

   for row in meters_json:
       print(row)
      
      url_req = requests.post(url=url, headers=headers, data=row) # SINGLE POST REQUEST PER JSON
        
      data = json.loads(row)
      
      # PRINT RESPONSE
      url_response = url_req.text
      if url_response == '{ }':           
           txt.insert(INSERT, '\nOK: Meter ' + data['tne_number'] + ' added to the Platform')
      else:
           txt.insert(INSERT, '\nERROR: Meter ' + data['tne_number'] + ' not added to the Platform')
           print(url_response)

#      print('\n' + meters_json[row])
      bar["value"]= current_bar_value
      bar.update()
      time.sleep(0.5)
      current_bar_value = current_bar_value + 1

     


if __name__ == "__main__":
   start()
